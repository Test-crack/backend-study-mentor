"""
Converts a verification report .xlsx (Layer 1 or Layer 2, drills or
diagnostic — same colored Summary + per-file-sheet shape either way) into a
PDF, so it can be read without Excel installed.

Usage:
    python scripts/seeders-ai/xlsxToPdf.py <input.xlsx> [output.pdf]

If output.pdf is omitted, writes next to the input with a .pdf extension.
One sheet -> one section in the PDF, in the same left-to-right order as the
workbook's tabs. Each cell keeps its fill colour (green/amber/red/grey) so a
PASS/FAIL/TOO_EASY row is exactly as visually obvious as it is in Excel.
"""
import sys
import os
from openpyxl import load_workbook
from reportlab.lib.pagesizes import landscape, A3, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT


def argb_to_hex(argb):
    """openpyxl fill colors are 'FFRRGGBB' (with an alpha byte) or None/theme-based."""
    if not argb or not isinstance(argb, str) or len(argb) < 6:
        return None
    rgb = argb[-6:]
    try:
        return colors.HexColor(f'#{rgb}')
    except Exception:
        return None


def cell_bg(cell):
    fill = cell.fill
    if fill is None or fill.fgColor is None:
        return None
    return argb_to_hex(fill.fgColor.rgb)


def cell_font_color(cell):
    font = cell.font
    if font is None or font.color is None:
        return None
    return argb_to_hex(font.color.rgb)


def col_width_for(sheet, col_idx, max_chars_cap=60):
    dim = sheet.column_dimensions.get(chr(64 + col_idx) if col_idx <= 26 else None)
    if dim and dim.width:
        return min(dim.width, max_chars_cap)
    return 15


def build_pdf(xlsx_path, pdf_path):
    wb = load_workbook(xlsx_path, data_only=True)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title2', parent=styles['Heading1'], fontSize=16, spaceAfter=6)
    sheet_title_style = ParagraphStyle('SheetTitle', parent=styles['Heading2'], fontSize=13, spaceBefore=4, spaceAfter=6, textColor=colors.HexColor('#1F3864'))
    cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=7.2, leading=9)
    plain_style = ParagraphStyle('Plain', parent=styles['Normal'], fontSize=9.5, leading=13)

    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=landscape(A3),
        topMargin=0.4 * inch, bottomMargin=0.4 * inch,
        leftMargin=0.3 * inch, rightMargin=0.3 * inch,
    )

    story = []
    story.append(Paragraph(os.path.basename(xlsx_path), title_style))
    story.append(Spacer(1, 6))

    for sheet_idx, sheet in enumerate(wb.worksheets):
        if sheet_idx > 0:
            story.append(PageBreak())
        story.append(Paragraph(sheet.title, sheet_title_style))

        rows = list(sheet.iter_rows())
        if not rows:
            story.append(Paragraph('(empty sheet)', plain_style))
            continue

        # Determine the widest row so every row in the table has the same column count.
        max_cols = max(len(r) for r in rows)

        table_data = []
        cell_colors = {}   # (row, col) -> background Color
        font_colors = {}   # (row, col) -> font Color
        span_rows = []      # row indices that are a single "wide text" row (e.g. title/notice rows)

        for r_idx, row in enumerate(rows):
            row_values = []
            is_wide_row = len(row) <= 2 and row[0].value is not None and (len(row) == 1 or row[1].value is None)
            for c_idx in range(max_cols):
                cell = row[c_idx] if c_idx < len(row) else None
                value = '' if cell is None or cell.value is None else str(cell.value)
                # Wrap long text in a Paragraph so it doesn't blow out the page width.
                para = Paragraph(value.replace('\n', '<br/>'), cell_style) if value else ''
                row_values.append(para)
                if cell is not None:
                    bg = cell_bg(cell)
                    if bg:
                        cell_colors[(r_idx, c_idx)] = bg
                    fc = cell_font_color(cell)
                    if fc:
                        font_colors[(r_idx, c_idx)] = fc
            table_data.append(row_values)
            if is_wide_row:
                span_rows.append(r_idx)

        # Column widths: scale relative to configured Excel widths, capped so the
        # table fits landscape A3. Falls back to equal widths if nothing configured.
        usable_width = landscape(A3)[0] - 0.6 * inch
        raw_widths = []
        for c_idx in range(max_cols):
            letter = chr(65 + c_idx) if c_idx < 26 else 'A'
            dim = sheet.column_dimensions.get(letter)
            raw_widths.append(dim.width if (dim and dim.width) else 12)
        total = sum(raw_widths) or 1
        col_widths = [max(0.35 * inch, usable_width * (w / total)) for w in raw_widths]

        table = Table(table_data, colWidths=col_widths, repeatRows=0)

        style_cmds = [
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cccccc')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ]
        for (r, c), bg in cell_colors.items():
            style_cmds.append(('BACKGROUND', (c, r), (c, r), bg))
        for r_idx in span_rows:
            style_cmds.append(('SPAN', (0, r_idx), (max_cols - 1, r_idx)))

        table.setStyle(TableStyle(style_cmds))
        story.append(table)
        story.append(Spacer(1, 10))

    doc.build(story)


def main():
    if len(sys.argv) < 2:
        print('Usage: python xlsxToPdf.py <input.xlsx> [output.pdf]')
        sys.exit(1)

    xlsx_path = sys.argv[1]
    pdf_path = sys.argv[2] if len(sys.argv) > 2 else os.path.splitext(xlsx_path)[0] + '.pdf'

    build_pdf(xlsx_path, pdf_path)
    print(f'Wrote {pdf_path}')


if __name__ == '__main__':
    main()
